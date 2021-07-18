from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['B1'] = 2020
ws['C1'] = 2019
ws['D1'] = 2018
ws['E1'] = 2017
ws['F1'] = 2016
def Total_net_sales(need):
	if need:
		TNS2020 = input('請輸入2020淨銷售額:') # Rows can also be appended
		TNS2019 = input('請輸入2019淨銷售額:')
		TNS2018 = input('請輸入2018淨銷售額:')
		TNS2017 = input('請輸入2017淨銷售額:')
		TNS2016 = input('請輸入2016淨銷售額:')
		ws.append(["Total net sales", TNS2020, TNS2019, TNS2018, TNS2017, TNS2016 ])
def Net_income(need):
	if need:
		Ni2020 = input('請輸入2020淨利:') # Rows can also be appended
		Ni2019 = input('請輸入2019淨利:')
		Ni2018 = input('請輸入2018淨利:')
		Ni2017 = input('請輸入2017淨利:')
		Ni2016 = input('請輸入2016淨利:')
		ws.append(["Net income", Ni2020, Ni2019, Ni2018, Ni2017, Ni2016 ])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()


Total_net_sales(False)
Net_income(True)
# Save the file
wb.save("sample.xlsx")